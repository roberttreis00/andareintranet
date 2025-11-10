from django.utils.datastructures import MultiValueDictKeyError
from .forms import (SugestaoCompras, SugestaoComprasProgramada, GetSugestaoCompras, FiltroDataForm, ConsultarCusto,
                    AtualizarCusto, FiltroPeriodoAnterior, FiltroLucroLiquido)
from django.urls import reverse_lazy
from django.views.generic.edit import FormView
from . import functions_compras
from django.views import View
from django.shortcuts import get_object_or_404
import os
from django.http import JsonResponse

# from .functions_analise_dados import saldo_estoque
from .tasks import atualizar_lista_produtos, tratar_sugestao_de_compras
from .models import ProdutosCadastradosTiny, ArquivosProcessados, ProdutosAtivosTiny, DataUltimaAtualizacaoCustos, \
    Pedidos
from datetime import datetime, timezone
from django.http import HttpResponse
import openpyxl
from datetime import date, timedelta
from . import functions_analise_dados
from django.contrib import messages
from django.utils import timezone as ti
from django.db.models import Max
from dateutil.relativedelta import relativedelta


class GerarSugestaoCompras(FormView):
    template_name = 'sugestao.html'
    form_class = SugestaoCompras
    success_url = reverse_lazy("gerar-sugestao-de-compras")

    def form_valid(self, form):
        # Aqui posso tratar os dados enviados
        workbook_estoque_full = self.request.FILES['PlanilhaEstoqueFull']
        workbook_estoque_tiny = self.request.FILES['PlanilhaSaldoTiny']
        workbook_relatorio_vendas = self.request.FILES['PlanilhaRelatorioDeVendas']
        try:
            workbook_ordens_de_compra = self.request.FILES['PlanilhaOrdemDeCompras']
        except MultiValueDictKeyError:
            workbook_ordens_de_compra = None
        marca_giro = form.cleaned_data['Marca']
        periodo = form.cleaned_data['Periodo']

        # Ao ser enviado os dados para gerar o giro de compras acionar a função que atualizar as marcas no banco
        functions_compras.verificar_relatorio_vendas_marca_nova(workbook_relatorio_vendas)
        # E a tarefa para atualizar a lista de SKUS Ativos e Inativos/ Uma tarefa que demora precisa colocar na fila do Redis
        ultima_atualizacao_produtos = (ProdutosCadastradosTiny.objects.get(Nome_Lista_Produtos='Produtos Ativos')
                                       .Data_ultima_atualizacao)
        data_atual = datetime.now(timezone.utc)
        qtd_dias_sem_atualizar = data_atual - ultima_atualizacao_produtos
        # Antes verificar se já faz mais de 14 Dias que não tem atualização
        if qtd_dias_sem_atualizar.days > 14:
            atualizar_lista_produtos.delay()  # Mada para fila a tarefa de atualizar o sistema de skus

        workbook_relatorio_vendas.seek(0)

        # Segue o fluxo normal agora trata os dados das planilhas e return a Sugestão de Compras
        df_sugestao_compras_win_and_loser = functions_compras.gerar_sugestao_compras(
            workbook_estoque_full, workbook_estoque_tiny, workbook_relatorio_vendas, workbook_ordens_de_compra,
            marca_giro, periodo)

        df_win = df_sugestao_compras_win_and_loser[0]
        df_loser = df_sugestao_compras_win_and_loser[1]

        # Com a sugestao de compras tratar dados para retornar um response e assim realizar o Download
        # Acrescentar duas colunas df_win
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Produtos Giro OK'
        ws.append(list(df_win.columns) + ['Desvio'])

        for idx, row in df_win.iterrows():
            ws.append(
                [row["SKU_Seller"], row["Estoque Geral"], row["Estoque Full"], row["Estoque Comprado"],
                 row["Saídas Periodo"], row["Sugestão de Compras"], row['Ajuste Comprador'],
                 ""])  # Duas colunas extras vazias

        for row_idx in range(2, len(df_win) + 2):  # Começa na linha 2 (1 é cabeçalho)
            ws[f"H{row_idx}"].value = (
                f'=IF(G{row_idx}="","",'
                f'IF(G{row_idx}=F{row_idx},"",'
                f'IF(AND(F{row_idx}<0,G{row_idx}=0),"",'
                f'IF(G{row_idx}>F{row_idx},"ALTERADO",IF(G{row_idx}<F{row_idx},"ALTERADO","")))))'
            )

        # Acrescentar duas colunas df_loser
        ws2 = wb.create_sheet(title='Produtos Giro Baixo')
        ws2.append(list(df_loser.columns) + ['Desvio'])

        for idx, row in df_loser.iterrows():
            ws2.append(
                [row["SKU_Seller"], row["Estoque Geral"], row["Estoque Full"], row["Estoque Comprado"],
                 row["Saídas Periodo"], row["Sugestão de Compras"], row['Ajuste Comprador'],
                 ""])  # Duas colunas extras vazias

        for row_idx in range(2, len(df_loser) + 2):  # Começa na linha 2 (1 é cabeçalho)
            ws2[f"H{row_idx}"].value = (
                f'=IF(G{row_idx}="","",'
                f'IF(G{row_idx}=F{row_idx},"",'
                f'IF(AND(F{row_idx}<0,G{row_idx}=0),"",'
                f'IF(G{row_idx}>F{row_idx},"ALTERADO",IF(G{row_idx}<F{row_idx},"ALTERADO","")))))'
            )
        # Agora juntar os dois DF em um só
        # Return response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="sugestão_compras_{marca_giro}.xlsx"'
        wb.save(response)
        return response


# Gera a sugestão de compras programada é o mesmo que sugestão giro 1 2 3 meses só que tem 3 semestre de relatorio
# de vendas para cálculos
class GerarSugestaoProgramada(FormView):
    template_name = 'programada.html'
    form_class = SugestaoComprasProgramada
    success_url = reverse_lazy("gerar-sugestao-programada")

    # Agora tratar os dados fazer o mesmo só que agora com duas colunas a mais s
    def form_valid(self, form):
        workbook_estoque_full = self.request.FILES['PlanilhaEstoqueFull']
        workbook_estoque_tiny = self.request.FILES['PlanilhaSaldoTiny']
        workbook_relatorio_1 = self.request.FILES['PlanilhaRelatorioDeVendas1']
        workbook_relatorio_2 = self.request.FILES['PlanilhaRelatorioDeVendas2']
        workbook_relatorio_3 = self.request.FILES['PlanilhaRelatorioDeVendasSemestreAtual']

        try:
            workbook_ordens_de_compra = self.request.FILES['PlanilhaOrdemDeCompras']
        except MultiValueDictKeyError:
            workbook_ordens_de_compra = None
        marca_giro = form.cleaned_data['Marca']

        # Segue o fluxo normal agora trata os dados das planilhas e return a Sugestão de Compras Programada sem precisar
        # Acrescentar novas marcas caso tenha deixa para o Giro
        df_compras_crescimento = functions_compras.gerar_sugestao_compras_programada(workbook_estoque_full,
                                                                                     workbook_estoque_tiny,
                                                                                     workbook_relatorio_1,
                                                                                     workbook_relatorio_2,
                                                                                     workbook_relatorio_3,
                                                                                     workbook_ordens_de_compra,
                                                                                     marca_giro)

        df_compras = df_compras_crescimento[0]
        crescimento = df_compras_crescimento[1]

        # Agora com o DataFrame em mãos acrescentar as colunas e retorna o giro programado
        # Acrescentar duas colunas df_win
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sugestão de Compras Semestral'
        ws.append(list(df_compras.columns) + ['Desvio'])

        for idx, row in df_compras.iterrows():
            ws.append(
                [row["SKU_Seller"], row["Estoque Geral"], row["Estoque Full"], row["Estoque Comprado"],
                 row["Saídas 1° Semestre"], row['Saídas 2° Semestre'], row['Saídas Semestre Atual'],
                 row["Sugestão de Compras"], row['Ajuste Comprador'], ""])  # Duas colunas extras vazias

        for row_idx in range(2, len(df_compras) + 2):  # Começa na linha 2 (1 é cabeçalho)
            ws[f"J{row_idx}"].value = (
                f'=IF(I{row_idx}="","",'
                f'IF(I{row_idx}=G{row_idx},"",'
                f'IF(AND(H{row_idx}<0,I{row_idx}=0),"",'
                f'IF(I{row_idx}>H{row_idx},"ALTERADO",IF(I{row_idx}<H{row_idx},"ALTERADO","")))))'
            )

        # Agora juntar os dois DF em um só
        # Return response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="sugestao_compras_{marca_giro}_{crescimento}%.xlsx"'
        wb.save(response)
        return response


class GerarOrdemDeCompra(FormView):
    template_name = 'ordem_compras.html'
    form_class = GetSugestaoCompras
    success_url = reverse_lazy('gerar-ordem-de-compra-tiny')

    def form_valid(self, form):
        upload_file = form.cleaned_data['PlanilhaSugestaoCompras']
        fornecedor = form.cleaned_data['Fornecedor']
        situacao = form.cleaned_data['Situacao_compra']

        # 1. Criar uma instancia no banco
        # 2. Chama a tarefa com tarefa.delay(task_instance.id)
        task_instance = ArquivosProcessados.objects.create(
            Workbook=upload_file,
            Fornecedor=fornecedor,
            situacao_compra=situacao,  # situação giro ou programação
            status='Pendente',
        )

        # disparar a tarefa no Celery
        tratar_sugestao_de_compras.delay(task_instance.id)

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Exibe as últimas 10 tarefas para o usuário
        context['tasks'] = ArquivosProcessados.objects.order_by('-created_at')[:1]
        return context


class TaskStatusView(View):
    def get(self, request, task_id):
        task = get_object_or_404(ArquivosProcessados, id=task_id)
        data = {
            'status': task.status,
            'output_file_url': task.output_file.url if task.output_file else None
        }
        return JsonResponse(data)


class DownloadFileView(View):

    def get(self, request, task_id):
        task = get_object_or_404(ArquivosProcessados, id=task_id)
        if task.status == 'Completo' and task.output_file:
            file_path = task.output_file.path
            if os.path.exists(file_path):
                with open(file_path, 'rb') as fh:
                    response = HttpResponse(fh.read(),
                                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")  # MIME type para .xlsx
                    response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(file_path)
                    return response
        return HttpResponse("Arquivo não encontrado ou tarefa não concluída.", status=404)


class DashboardAndare(FormView):
    template_name = 'dashboard.html'
    form_class = FiltroDataForm
    success_url = reverse_lazy('dashboard')

    def get_initial(self):
        """Preenche o formulário com datas padrão (últimos 30 dias)"""
        initial = super().get_initial()
        initial['data_inicio'] = date.today() - timedelta(days=30)
        initial['data_fim'] = date.today()
        return initial

    def form_valid(self, form):
        # Pega os dados do formulário ou usa padrão
        data_inicio = form.cleaned_data.get("data_inicio") or (date.today() - timedelta(days=30))
        data_fim = form.cleaned_data.get("data_fim") or date.today()
        marca = form.cleaned_data.get('Marca')

        # Agora podemos calcular os dados
        qtd_pedidos = functions_analise_dados.quantidade_vendas_do_periodo(data_inicio, data_fim, marca)
        faturamento = functions_analise_dados.faturamento_total(data_inicio, data_fim, marca)
        ticket_medio = round(faturamento / qtd_pedidos, 2) if qtd_pedidos else 0
        faturamento_mkt = functions_analise_dados.faturamento_por_marketplace(data_inicio, data_fim, marca)
        top5_skus_mais_vendidos = functions_analise_dados.skus_mais_vendido(data_inicio, data_fim, marca)
        valor_total_frete = functions_analise_dados.custo_frete_total(data_inicio, data_fim, marca)

        context = self.get_context_data(
            form=form,
            data_inicio=data_inicio,
            data_fim=data_fim,
            qtd_pedidos=qtd_pedidos,
            faturamento=f"{faturamento:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            ticket_medio=ticket_medio,
            marketplace=list(faturamento_mkt.keys()),
            qtd_por_mkt=list(faturamento_mkt.values()),
            top5_sku_vendidos=list(top5_skus_mais_vendidos.keys()),
            qtd_top5=list(top5_skus_mais_vendidos.values()),
            frete = valor_total_frete,
            valor_total = f"{faturamento+valor_total_frete:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
        )
        return self.render_to_response(context)


class CurvaABC(FormView):
    template_name = 'curva_abc.html'
    form_class = FiltroDataForm
    success_url = reverse_lazy("curva_abc")

    def get_initial(self):
        """Preenche o formulário com datas padrão (últimos 30 dias)"""
        initial = super().get_initial()
        initial['data_inicio'] = date.today() - timedelta(days=30)
        initial['data_fim'] = date.today()
        return initial

    def form_valid(self, form):
        data_inicio = form.cleaned_data.get("data_inicio") or (date.today() - timedelta(days=30))
        data_fim = form.cleaned_data.get("data_fim") or date.today()
        marca = form.cleaned_data.get('Marca')

        curva_abc_raw = functions_analise_dados.curva_abc(data_inicio, data_fim, marca)

        # Tentar obter o saldo da planilha, se enviada
        try:
            workbook_estoque_tiny = self.request.FILES['PlanilhaSaldoTiny']
            saldo_skus_pai = functions_analise_dados.saldo_estoque(workbook_estoque_tiny)
        except MultiValueDictKeyError:
            saldo_skus_pai = {}

        # Preparar curva_abc processado: cada categoria com lista de (sku, valor, saldo)
        curva_abc = {}
        for categoria, data in curva_abc_raw[0].items():
            skus_5 = list(data['skus'].items())
            skus_valor_saldo_qtd = [
                (sku, valor, saldo_skus_pai.get(sku, 0), curva_abc_raw[1].get(sku, 0))
                for sku, valor in skus_5
            ]

            # Guarda lista e total
            curva_abc[categoria] = {
                'skus_valor_saldo': skus_valor_saldo_qtd,
                'total': data.get('total', 0),
                'rowspan': len(skus_valor_saldo_qtd)
            }

        # Preparar labels e valores para gráfico
        labels = list(curva_abc.keys())
        valores = [round(d["total"], 2) for d in curva_abc.values()]

        context = self.get_context_data(
            form=form,
            curva_abc=curva_abc,
            curva_abc_labels=labels,
            curva_abc_valores=valores,
        )
        return self.render_to_response(context)


class Custos(FormView):
    form_class = ConsultarCusto
    template_name = 'custos.html'
    success_url = reverse_lazy("custos")

    def form_valid(self, form):
        sku_ean = form.cleaned_data['sku_ean_pesquisado']
        try:
            try:
                # Quer dizer que é por EAN
                int(sku_ean)
                produto = ProdutosAtivosTiny.objects.get(ean=sku_ean)
                custo = produto.custo
            except ValueError:
                # Quer dizer que é por SKU
                produto = ProdutosAtivosTiny.objects.get(sku=sku_ean.upper())
                custo = produto.custo
        except ProdutosAtivosTiny.DoesNotExist:
            produto = False

        if produto:
            return self.render_to_response({'form': form, 'resultado': custo})
        else:
            return self.render_to_response({'form': form, 'resultado': 'Não encontrado'})


class AtualizarCustos(FormView):
    template_name = "atualizar_custos.html"
    form_class = AtualizarCusto
    success_url = reverse_lazy("atualizar_custos")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ultima_data = DataUltimaAtualizacaoCustos.objects.order_by('-DataUltima').first()
        context['ultima_data'] = ultima_data.DataUltima if ultima_data else None
        return context

    def form_valid(self, form):
        arquivo_zipado = self.request.FILES['arquivo_zip_nfs']
        functions_analise_dados.atualizar_custos_produtos(arquivo_zipado)

        messages.success(self.request, "Custos atualizados com sucesso!")

        ultima_data = DataUltimaAtualizacaoCustos.objects.order_by('-DataUltima').first()
        ultima_data.DataUltima = ti.now().date()
        ultima_data.save()

        return super().form_valid(form)


class PeriodoAnterior(FormView):
    template_name = 'periodo_anterior.html'
    form_class = FiltroPeriodoAnterior
    success_url = reverse_lazy('periodo_anterior')

    def form_valid(self, form):
        marca = form.cleaned_data.get('Marca')
        periodo = form.cleaned_data.get('Periodo')

        periodo = int(periodo)
        # Pegar a ultima data do ultimo resgistro feito de pedido do banco de dados
        mais_atual = Pedidos.objects.aggregate(Max('data_pedido'))['data_pedido__max']
        data_inicio = mais_atual - relativedelta(days=periodo)

        # Periodo Atual
        periodo_atual_qtd = functions_analise_dados.quantidade_vendas_do_periodo(data_inicio, mais_atual, marca)
        periodo_atual_fat = functions_analise_dados.faturamento_total(data_inicio, mais_atual, marca)
        # print(data_inicio, mais_atual)
        # print(f"Vendeu: {periodo_atual_qtd} Unds Faturou: R${periodo_atual_fat}")

        datafim2 = data_inicio - relativedelta(days=1)
        data_inicio2 = data_inicio - relativedelta(days=periodo)

        # Period Anterior
        periodo_anterior_qtd = functions_analise_dados.quantidade_vendas_do_periodo(data_inicio2, datafim2, marca)
        periodo_anterior_fat = functions_analise_dados.faturamento_total(data_inicio2, datafim2, marca)
        # print(f"Vendeu: {periodo_anterior_qtd} Unds Faturou: R${periodo_anterior_fat}")

        # print(data_inicio2, datafim2)

        crescimento_qtd = round(((periodo_atual_qtd - periodo_anterior_qtd) / periodo_atual_qtd) * 100, 2)
        crescimento_fat = round(((periodo_atual_fat - periodo_anterior_fat) / periodo_atual_fat) * 100, 2)

        # print(crescimento_fat, crescimento_qtd)
        data_anterior_1 = datetime(data_inicio2.year, data_inicio2.month, data_inicio2.day).strftime("%d/%m/%Y")
        data_anterior_2 = datetime(data_inicio.year, data_inicio.month, data_inicio.day).strftime("%d/%m/%Y")
        data_atual_1 = datetime(mais_atual.year, mais_atual.month, mais_atual.day).strftime("%d/%m/%Y")
        data_atual_2 = datetime(datafim2.year, datafim2.month, datafim2.day).strftime("%d/%m/%Y")

        print(data_anterior_1, data_anterior_2)
        print(data_atual_2, data_atual_1)

        context = self.get_context_data(
            form=form,
            qtd_ant = periodo_anterior_qtd,
            fat_ant = periodo_anterior_fat,
            qtd_atu = periodo_atual_qtd,
            fat_atu = periodo_atual_fat,
            por_qtd = crescimento_qtd,
            por_fat = crescimento_fat,
            dat_ant = data_anterior_1,
            data_3 = data_atual_1,
        )
        return self.render_to_response(context)

import locale
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

class LucroLiquido(FormView):
    template_name = 'lucro_liquido.html'
    form_class = FiltroLucroLiquido
    success_url = reverse_lazy('lucro_liquido')

    def form_valid(self, form):
        data_inicio = form.cleaned_data.get("data_inicio")
        data_fim = form.cleaned_data.get("data_fim")
        taxa_mkt = form.cleaned_data.get("taxa_marketplace")
        taxa_fixa = form.cleaned_data.get("taxa_fixa")
        mkt = form.cleaned_data.get("marketplace")
        # print(data_inicio, data_fim, taxa_mkt, taxa_fixa, mkt)
        data = functions_analise_dados.calcular_lucro_liquido(data_inicio, data_fim, taxa_mkt, taxa_fixa, mkt)

        porcentagem_lucro = round((data[0] / data[1]) * 100, 2)

        context = self.get_context_data(
            form=form,
            faturamento=locale.format_string('%.2f', data[1], grouping=True),
            lucro_liquido=locale.format_string('%.2f', data[0], grouping=True),
            lucro_por_cem = porcentagem_lucro
        )

        return self.render_to_response(context)
